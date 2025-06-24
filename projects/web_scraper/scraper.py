import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
import re
import logging
from time import sleep
from typing import List, Dict, Optional
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('scraper.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def clean_price(price_str: str) -> float:
    """Convert price strings like '£51.77' to float (51.77)

    Args:
        price_str: Price string containing currency symbol and value

    Returns:
        Cleaned price as float. Returns 0.0 if conversion fails.
    """
    try:
        # Remove all non-digit characters except decimal point
        return float(re.sub(r'[^\d.]', '', price_str))
    except (TypeError, ValueError) as e:
        logger.error(f"Price cleaning error: {price_str} | {e}")
        return 0.0


def scrape_page(url: str) -> List[BeautifulSoup]:
    """Scrape a single page and extract book elements

    Args:
        url: URL of the page to scrape

    Returns:
        List of book elements found on the page
    """
    try:
        response = requests.get(
            url,
            headers={'User-Agent': 'Mozilla/5.0'},
            timeout=15
        )
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')
        return soup.find_all('article', class_='product_pod')

    except requests.exceptions.RequestException as e:
        logger.error(f"Request failed: {url} | {e}")
        return []
    except Exception as e:
        logger.error(f"Unexpected error: {url} | {e}")
        return []


def parse_book(book: BeautifulSoup, page: int) -> Optional[Dict]:
    """Extract data from a single book element

    Args:
        book: BeautifulSoup element containing book data
        page: Page number where the book was found

    Returns:
        Dictionary with book data or None if parsing fails
    """
    try:
        return {
            'Title': book.h3.a['title'],
            'Price (£)': clean_price(book.find('p', class_='price_color').text),
            'Rating': book.p['class'][1],
            'Stock': book.find('p', class_='instock').text.strip(),
            'Page': page
        }
    except Exception as e:
        logger.error(f"Book parsing error | {e}")
        return None


def scrape_books(base_url: str, max_pages: int = 3, delay: float = 1.0) -> pd.DataFrame:
    """Scrape multiple pages of books

    Args:
        base_url: Base URL with {page} placeholder
        max_pages: Maximum number of pages to scrape
        delay: Delay between page requests (seconds)

    Returns:
        DataFrame containing all scraped books
    """
    all_books = []

    for page in range(1, max_pages + 1):
        logger.info(f"Scraping page {page}/{max_pages}")
        page_url = base_url.format(page)
        books = scrape_page(page_url)

        if not books:
            logger.warning(f"No books found on page {page}")
            continue

        for book in books:
            book_data = parse_book(book, page)
            if book_data:
                all_books.append(book_data)

        # Respectful scraping delay
        sleep(delay)

    return pd.DataFrame(all_books)


def save_to_excel(df: pd.DataFrame, output_dir: str = 'output') -> str:
    """Save DataFrame to Excel with formatting using openpyxl

    Args:
        df: DataFrame containing book data
        output_dir: Output directory path

    Returns:
        Path to the created Excel file
    """
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, 'books_data.xlsx')

    try:
        # Save main data
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Books')

            # Create summary sheet
            summary = df.describe(include='all').fillna('')
            summary.to_excel(writer, sheet_name='Summary')

            # Access workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Books']

            # Header styling
            header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
            header_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Apply styling to headers
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            # Auto-adjust column widths
            for col_idx, column in enumerate(df.columns, 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)

                # Find max length in column
                for cell in worksheet[col_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                # Set column width with buffer
                adjusted_width = max_length + 2
                worksheet.column_dimensions[col_letter].width = adjusted_width

        logger.info(f"Data saved to {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"Excel creation failed: {e}")
        raise


def main() -> None:
    """Main execution function"""
    try:
        BASE_URL = "http://books.toscrape.com/catalogue/page-{}.html"

        logger.info("Starting book scraping process")
        df = scrape_books(BASE_URL, max_pages=3, delay=1.0)

        if not df.empty:
            output_path = save_to_excel(df)
            logger.info(f"✅ Success! Total books scraped: {len(df)}")
            logger.info(f"📊 Average price: £{df['Price (£)'].mean():.2f}")
            logger.info(f"💾 Output file: {output_path}")
        else:
            logger.error("❌ No data scraped. Exiting program.")

    except Exception as e:
        logger.exception("🛑 Critical error in main execution")


if __name__ == "__main__":
    main()